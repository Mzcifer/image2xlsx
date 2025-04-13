from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import argparse

def image_to_excel(image_path, output_file, cell_size=2, row_height=14, max_size=100):
    """
    将图片转换为Excel彩色表格
    参数：
    image_path: 输入图片路径
    output_file: 输出Excel文件路径
    cell_size: 单元格宽度（字符单位）
    row_height: 行高（点数）
    max_size: 最大行列尺寸（保持长宽比）
    """
    # 加载并预处理图片
    img = Image.open(image_path)
    img = img.convert("RGB")
    
    # 等比例缩放图片
    ratio = min(max_size/img.width, max_size/img.height)
    new_size = (int(img.width*ratio), int(img.height*ratio))
    img = img.resize(new_size, Image.Resampling.LANCZOS)

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    
    # 设置单元格尺寸
    for col in range(1, new_size[0]+1):
        ws.column_dimensions[get_column_letter(col)].width = cell_size
    for row in range(1, new_size[1]+1):
        ws.row_dimensions[row].height = row_height

    # 遍历每个像素填充颜色
    for y in range(new_size[1]):
        for x in range(new_size[0]):
            # 获取像素颜色（注意PIL的坐标顺序）
            r, g, b = img.getpixel((x, y))
            
            # 转换为Excel颜色格式（AARRGGBB）
            hex_color = f"FF{r:02x}{g:02x}{b:02x}".upper()
            
            # 创建填充样式
            fill = PatternFill(fill_type="solid", fgColor=hex_color)
            
            # 设置单元格（注意行列对应关系）
            cell = ws.cell(row=y+1, column=x+1)
            cell.fill = fill

        # 打印进度
        print(f"处理进度: {y+1}/{new_size[1]}行")

    # 保存文件
    wb.save(output_file)
    print(f"文件已保存至：{output_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="将图片转换为Excel彩色表格")
    parser.add_argument("input", help="输入图片路径")
    parser.add_argument("-o", "--output", default="output.xlsx", help="输出文件路径")
    parser.add_argument("-s", "--size", type=int, default=100, help="最大边长尺寸")
    args = parser.parse_args()

    image_to_excel(args.input, args.output, max_size=args.size)
